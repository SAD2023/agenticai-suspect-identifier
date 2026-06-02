"""
create_sample_template.py — Input Template Generator
======================================================
Generates a blank Excel template that analysts fill in with the individuals
they want to score each day.

"""

import os
import logging
import pandas as pd

from config import FEATURE_CONFIG, FILE_CONFIG

logger = logging.getLogger(__name__)


def generate_template(output_path: str = None) -> str:
    """
    Args:
        output_path (str): Destination .xlsx file path.
                           Defaults to outputs/input_template.xlsx.

    Returns:
        str: The path where the template was saved.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required to generate the template.\n"
            "Install it with:  pip install openpyxl"
        )

    if output_path is None:
        output_path = os.path.join(FILE_CONFIG["DEFAULT_OUTPUT_DIR"], "input_template.xlsx")

    # ---- Determine column structure ----
    feature_columns = FEATURE_CONFIG["FEATURE_COLUMNS"]
    id_col          = FEATURE_CONFIG["ID_COLUMN"]
    name_col        = FEATURE_CONFIG["NAME_COLUMN"]

    using_placeholders = not bool(feature_columns)
    if using_placeholders:
        # 
        feature_columns = [
            "watchlist_match_score",
         "prior_security_incidents",
        "trips_to_risk_zones",
        "last_minute_booking_rate",
        "document_anomaly_score",
        "network_association_score",
        "financial_activity_flag",
        "cross_agency_flag_count",
        "border_screening_anomaly_score",
        "encrypted_comms_flag",`
        ]
        logger.warning(
            "FEATURE_COLUMNS is empty in config.py. "
            
        )

 
    all_columns = [id_col]
    if name_col:
        all_columns.append(name_col)
    all_columns.extend(feature_columns)

    wb = Workbook()


    ws_instructions = wb.active
    ws_instructions.title = "INSTRUCTIONS"
    _write_instructions_sheet(ws_instructions, all_columns, using_placeholders)


    ws_data = wb.create_sheet("Person List")
    _write_data_sheet(ws_data, all_columns, id_col, name_col, feature_columns)

    wb.save(output_path)
    logger.info(f"Input template saved → {output_path}")
    return output_path


def _write_instructions_sheet(ws, all_columns: list, using_placeholders: bool) -> None:
    from openpyxl.styles import Font, Alignment

    title_font   = Font(bold=True, size=14)
    heading_font = Font(bold=True, size=11)
    body_font    = Font(size=10)
    warn_font    = Font(bold=True, size=11, color="CC0000")

    ws.column_dimensions["A"].width = 90

    def row(text="", font=None, wrap=True):
        ws.append([text])
        cell = ws[ws.max_row][0]
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=wrap)

    row("Suspect Prioritization System — Analyst Input Template", font=title_font)
    row()



    row("HOW TO USE THIS FILE", font=heading_font)
    row("1. Go to the 'Person List' sheet (tab at the bottom).")
    row("2. Add one row per individual you want to score today.")
    row(f"3. Fill in the '{all_columns[0]}' column with a unique identifier for each person.")
    if len(all_columns) > 1:
        row(f"4. Fill in all feature columns with the relevant data for each person.")
    row("5. Save the file.")
    row("6. Run:  python main.py predict --input <this_file.xlsx>")
    row("7. Open the output report in the 'outputs/' folder.")
    row()

    row("REQUIRED COLUMNS", font=heading_font)
    for col in all_columns:
        row(f"  • {col}")
    row()

    row("NOTES", font=heading_font)
    row("• Do not rename or reorder the column headers in 'Person List'.")
    row("• Leave cells blank (not zero) if a value is genuinely unknown.")
    row("• The 'name' column is for display only and is never used as a model input.")
    row("• Do not add columns not listed above — they will be ignored.")


def _write_data_sheet(
    ws,
    all_columns: list,
    id_col: str,
    name_col: str,
    feature_columns: list,
) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_COLOR = "2C3E7A"
    EXAMPLE_COLOR = "EEF2FF"

    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill("solid", fgColor=EXAMPLE_COLOR)

    ws.append(all_columns)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


    for i in range(1, 6):
        row_data = []
        for col in all_columns:
            if col == id_col:
                row_data.append(f"PERSON_{i:04d}")
            elif col == name_col:
                row_data.append(f"Example Name {i}")
            else:
                row_data.append(0.0) 
        ws.append(row_data)

        example_row = ws[ws.max_row]
        for cell in example_row:
            cell.fill = example_fill


    ws.append([])
    ws.append(["Replace example rows above with real data. "])
    note_cell = ws[ws.max_row][0]
    note_cell.font = Font(italic=True, color="888888")


    for col_idx, col_name in enumerate(all_columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 6, 18)

    ws.freeze_panes = "A2" 

if __name__ == "__main__":
    import sys
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    out = sys.argv[1] if len(sys.argv) > 1 else None
    path = generate_template(out)
    print(f"Template generated: {path}")
