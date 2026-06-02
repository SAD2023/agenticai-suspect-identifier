"""
report_generator.py — Output Report Generation
================================================
Produces formatted output files for analysts after a prediction run.

"""

import logging
import pandas as pd
from datetime import datetime

from config import FEATURE_CONFIG

logger = logging.getLogger(__name__)


TIER_FILL_COLORS = {
    "HIGH":   "FF6B6B",   # Red
    "MEDIUM": "FFC857",   # Amber
    "LOW":    "A8D8A8",   # Green
}

HEADER_COLOR = "2C3E7A"   # Dark blue for header rows
HEADER_FONT_COLOR = "FFFFFF"



def generate_excel_report(df: pd.DataFrame, output_path: str) -> None:
    """
    Writes a formatted, color-coded Excel workbook with three sheets.

    Args:
        df          (pd.DataFrame): Final scored, tiered, and explained DataFrame.
        output_path (str):          Destination .xlsx file path.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "openpyxl is required to generate Excel reports.\n"
            "Install it with:  pip install openpyxl"
        )

    wb = Workbook()

    # Sheet 1: Full prioritized list
    ws_full = wb.active
    ws_full.title = "Prioritized List"
    _write_sheet(ws_full, df, title="Prioritized List")

    # Sheet 2: HIGH priority only
    ws_high = wb.create_sheet("HIGH Priority")
    df_high = df[df["priority_tier"] == "HIGH"].copy()
    if not df_high.empty:
        _write_sheet(ws_high, df_high, title="HIGH Priority")
    else:
        ws_high.append(["No individuals were scored above the HIGH priority threshold."])

    # Sheet 3: Summary statistics
    ws_stats = wb.create_sheet("Summary Stats")
    _write_summary_sheet(ws_stats, df)

    wb.save(output_path)
    logger.info(f"Excel report saved → {output_path}")


def _write_sheet(ws, df: pd.DataFrame, title: str = "") -> None:
    """
    Writes a DataFrame to an openpyxl worksheet with formatted headers
    and row background coloring based on the 'priority_tier' column.

    Args:
        ws    : openpyxl Worksheet object.
        df    (pd.DataFrame): Data to write.
        title (str):          Optional title row prepended above the header.
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill      = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font      = Font(bold=True, color=HEADER_FONT_COLOR)
    header_alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    # Optional title row
    if title:
        ws.append([title])
        ws[1][0].font = Font(bold=True, size=13)
        ws.append([])   # Blank spacer row

    # Header row
    ws.append(list(df.columns))
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_alignment

    # Data rows with tier-based coloring
    for _, row_data in df.iterrows():
        ws.append(list(row_data))

        tier  = str(row_data.get("priority_tier", "LOW"))
        color = TIER_FILL_COLORS.get(tier, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)

        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-fit column widths (capped at 80 characters to avoid very wide columns)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_values = df[col_name].astype(str)
        max_content_len = col_values.str.len().max() if not df.empty else 0
        col_width = min(max(len(str(col_name)), max_content_len) + 4, 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Freeze the header row so it stays visible while scrolling
    freeze_row = header_row_idx + 1
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)


def _write_summary_sheet(ws, df: pd.DataFrame) -> None:
    """
    Writes aggregate statistics to the "Summary Stats" worksheet.
    Args:
        ws    : openpyxl Worksheet object.
        df    (pd.DataFrame): Scored + tiered DataFrame.
    """
    from openpyxl.styles import Font, PatternFill

    section_font  = Font(bold=True, size=11)
    header_fill   = PatternFill("solid", fgColor=HEADER_COLOR)
    header_font   = Font(bold=True, color=HEADER_FONT_COLOR)

    def section(title: str):
        ws.append([title])
        for cell in ws[ws.max_row]:
            cell.font = section_font
        ws.append([])

    ws.append(["Suspect Prioritization — Run Summary"])
    ws[1][0].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["Report generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Total individuals analyzed:", len(df)])
    ws.append([])

    section("Priority Tier Breakdown")
    ws.append(["Tier", "Count", "% of total"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    for tier in ["HIGH", "MEDIUM", "LOW"]:
        count = (df["priority_tier"] == tier).sum()
        pct   = 100.0 * count / len(df) if len(df) > 0 else 0.0
        row   = ws.append([tier, count, f"{pct:.1f}%"])
        fill  = PatternFill("solid", fgColor=TIER_FILL_COLORS.get(tier, "FFFFFF"))
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.append([])
    section("Risk Score Distribution")
    ws.append(["Statistic", "Value"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    scores = df["risk_score"]
    for label, val in [
        ("Minimum",        scores.min()),
        ("Maximum",        scores.max()),
        ("Mean (average)", scores.mean()),
        ("Median",         scores.median()),
    ]:
        ws.append([label, round(float(val), 4)])

    for col_letter in ("A", "B", "C"):
        ws.column_dimensions[col_letter].width = 30


def generate_pdf_summary(df: pd.DataFrame, output_path: str) -> None:
    """
    Generates a printable PDF summary report.

    Args:
        df          (pd.DataFrame): Final scored, tiered, explained DataFrame.
        output_path (str):          Destination .pdf file path.
    """
    try:
        from fpdf import FPDF
    except ImportError:
        logger.warning(
            "fpdf2 is not installed — falling back to plain-text output.\n"
            "Install fpdf2 with:  pip install fpdf2"
        )
        txt_path = output_path.replace(".pdf", "_report.txt")
        _generate_txt_fallback(df, txt_path)
        return

    id_col   = FEATURE_CONFIG.get("ID_COLUMN",   "person_id")
    name_col = FEATURE_CONFIG.get("NAME_COLUMN", "name")

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title block
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Suspect Prioritization Report", new_x="LMARGIN", new_y="NEXT", align="C")

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.cell(0, 7, f"Total individuals analyzed: {len(df)}",
             new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(5)

    # Tier breakdown summary box
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, "Priority Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 10)
    for tier in ["HIGH", "MEDIUM", "LOW"]:
        count = (df["priority_tier"] == tier).sum()
        pdf.cell(0, 6, f"  {tier}: {count} individual(s)", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # Detailed entries for HIGH and MEDIUM tiers
    for tier in ["HIGH", "MEDIUM"]:
        df_tier = df[df["priority_tier"] == tier]
        if df_tier.empty:
            continue

        # Tier header band
        r, g, b = _hex_to_rgb(TIER_FILL_COLORS[tier])
        pdf.set_fill_color(r, g, b)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(
            0, 9,
            f"  {tier} PRIORITY  —  {len(df_tier)} individual(s)",
            new_x="LMARGIN", new_y="NEXT", fill=True,
        )
        pdf.ln(2)

        for _, row in df_tier.iterrows():
            person_id   = row.get(id_col,   "N/A")
            person_name = row.get(name_col, "N/A") if name_col else "N/A"
            score       = row["risk_score"]
            explanation = row.get("explanation", "No explanation available.")

            pdf.set_font("Helvetica", "B", 10)
            pdf.cell(
                0, 7,
                f"ID: {person_id}  |  Name: {person_name}  |  Risk Score: {score:.3f}",
                new_x="LMARGIN", new_y="NEXT",
            )

            pdf.set_font("Helvetica", "", 9)
            pdf.multi_cell(0, 5, explanation)
            pdf.ln(2)

        pdf.ln(4)

    pdf.output(output_path)
    logger.info(f"PDF report saved → {output_path}")


def _generate_txt_fallback(df: pd.DataFrame, output_path: str) -> None:
    """
    Plain-text fallback report when fpdf2 is not installed.

    Args:
        df          (pd.DataFrame): Scored DataFrame.
        output_path (str):          Output .txt file path.
    """
    id_col   = FEATURE_CONFIG.get("ID_COLUMN",   "person_id")
    name_col = FEATURE_CONFIG.get("NAME_COLUMN", "name")

    lines = [
        "SUSPECT PRIORITIZATION REPORT",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Total individuals: {len(df)}",
        "=" * 70,
        "",
    ]
    for tier in ["HIGH", "MEDIUM", "LOW"]:
        df_tier = df[df["priority_tier"] == tier]
        lines.append(f"{tier} PRIORITY ({len(df_tier)} individuals)")
        lines.append("-" * 40)
        for _, row in df_tier.iterrows():
            lines.append(
                f"  ID: {row.get(id_col,'N/A')}  "
                f"Name: {row.get(name_col,'N/A') if name_col else 'N/A'}  "
                f"Score: {row['risk_score']:.3f}"
            )
            lines.append(f"  {row.get('explanation', '')}")
            lines.append("")
        lines.append("")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    logger.info(f"Text fallback report saved → {output_path}")


# ---------------------------------------------------------------------------
# Utility
# ---------------------------------------------------------------------------

def _hex_to_rgb(hex_color: str) -> tuple:
    """
    Converts a 6-character hex color string to an (R, G, B) integer tuple.

    Args:
        hex_color (str): e.g., "FF6B6B" (without leading #)

    Returns:
        tuple: (R, G, B) each in range 0–255.
    """
    h = hex_color.lstrip("#")
    return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
